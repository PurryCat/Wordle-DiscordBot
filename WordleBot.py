import os.path
import re
import random
import asyncio
from datetime import datetime
import discord
import pandas as pd
from discord.ext import commands
from discord.ext.commands import CommandNotFound
from discord.utils import get
from discord_components import DiscordComponents, Button
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime


bot = commands.Bot(command_prefix='')
DiscordComponents(bot)
bot.remove_command('help')

adminid = 259726664236269568
fixedTime = datetime.datetime(2022, 4, 30, 22, 30, 00)
token = ""

@bot.event
async def on_ready():
    if os.path.exists("wordle") == False:
        os.mkdir("wordle")
        wb = Workbook()
        wb.save(filename="wordle/bannedlist.xlsx")
        wb.save(filename="wordle/wordle.xlsx")
    if os.path.exists("players") == False:
        wb = Workbook()
        os.mkdir("players")
        
    if os.path.exists("wordle/words.xlsx") == False:
        print("Please place the words.xlsx file into the wordle folder")
        return
    print("Bot Online")

@bot.listen()
async def on_message(ctx):
    blacklist = [609884159590137927]
    if ctx.author.id in blacklist:
        return
    contentWords = ctx.content.split()
    try:
        contentWords[0] 
    except:
        return

    if ctx.author.bot: 
        return

# ------------------------------------------------------- Help -------------------------------------------------------
    elif contentWords[0].lower() == "fh" or contentWords[0].lower() == "fhelp" or contentWords[0].lower() == "fhelp":
        words = ctx.content.split()
        if len(words) > 1:
            if words[1] == "startwordle":
                embedVar = discord.Embed(title="Command Info: ``startwordle``", description="Shortcuts: ``sw`` \n Usage: ``fstartwordle`` \n \nStarts a game of wordle", color=0x00ff00)

            elif words[1] == "guess":
                embedVar = discord.Embed(title="Command Info: ``guess``", description="Usage: ``fguess`` ``word``  \n \nGuesses word for the command author's current game of wordle", color=0x00ff00)

            elif words[1] == "nolist":
                embedVar = discord.Embed(title="Command Info: ``nolist``", description="Shortcuts: ``nl`` \nUsage: ``fnolist`` \n \nDisplays the list of banned words and letters", color=0x00ff00)
            
            elif words[1] == "addword":
                embedVar = discord.Embed(title="Command Info: ``addword``", description="Shortcuts: ``aw`` \nUsage: ``faddword`` ``word`` \n \nAdds a word to the possible list of wordle words, has to be 5 characters and all english letters", color=0x00ff00)
            
            elif words[1] == "showgame":
                embedVar = discord.Embed(title="Command Info: ``showgame``", description="Shortcuts: ``sg`` \nUsage: ``fshowgame`` ``game number`` \n \nDisplays the specified game, ``fgames`` will show list of all possible games that user has played", color=0x00ff00)
            
            elif words[1] == "games":
                embedVar = discord.Embed(title="Command Info: ``games``", description="Usage: ``fgames`` ``page```` \n \nDisplays a list of the user's games, using ``fshowgame`` ``game`` will display that game of wordle", color=0x00ff00)

            elif words[1] == "userinfo":
                embedVar = discord.Embed(title="Command Info: ``userinfo``", description="Shortcuts: ``ui ``\nUsage: ``fuserinfo`` ``*user*`` \n \nDisplays user's statistics", color=0x00ff00)
                embedVar.set_footer(text= 'Optional parameters are shown in italics')
            
            elif words[1] == "board":
                embedVar = discord.Embed(title="Command Info: ``board``", description="Usage: ``fboard`` \n \nDisplays user's current game of wordle", color=0x00ff00)
                embedVar.set_footer(text= 'Optional parameters are shown in italics')

            elif words[1] == "forfeitwordle":
                embedVar = discord.Embed(title="Command Info: ``forfeitwordle``", description="Shortcuts: ``fff`` \nUsage: ``fff`` \n \nForfeits your current game of Wordle, resulting in a loss", color=0x00ff00)
        else:
            embedVar = discord.Embed(title="Wordle Help", description="Type ``fhelp command`` to see more details about a particular command.", color=0x00ff00)
            embedVar.add_field(name="Wordle" , value="``startwordle`` , ``guess``, ``nolist``, ``addword``, ``showgame``, ``games``, ``userinfo``, ``board``, ``forfeit``", inline=True)
        await ctx.channel.send(embed=embedVar)

# ------------------------------------------------------- Displays All Games The User Has Played -------------------------------------------------------
    if contentWords[0].lower() == "fgames":
        page  = re.findall(r'[0-9]+', ctx.content)
        workbook = load_workbook(filename="players/" + str(ctx.author.id) + "/wordles")
        sheet = workbook.active
        games = len(sheet['A'])
        if len(page) >= 1:
            page = page[0]
        else:
            page = 1
        if games == 1 and str(sheet.cell(row=1, column=1).value) == "None":
            await ctx.channel.send(str(ctx.author.mention) + " you havent played wordle before, do ``fstartwordle`` to start one")
        else:
            pages = ((games - (games%10)) / 10) + 1
            if int(page) > int(pages):
                await ctx.channel.send(str(ctx.author.mention) + " you dont have that many pages")
            else:
                if page == 1:
                    for x in range(10):
                        if sheet.cell(row=x+1, column=9).value == 0:
                            winorloss = "Loss"
                        elif sheet.cell(row=x+1, column=9).value == 1:
                            winorloss = "Win"
                        else:
                            winorloss = "Forfeit"
                        if x == 0:
                            if str(sheet.cell(row=x+1, column=8).value) != "None":
                                description = "**Game 1:** | " + "Word: **" + str(sheet.cell(row=x+1, column=8).value) + "** | Guesses: **" + str(sheet.cell(row=x+1, column=7).value) + "** | " + winorloss + "\n"
                            else:
                                await ctx.channel.send(str(ctx.author.mention) + " you havent played wordle before, do ``fstartwordle`` to start one")
                                return
                        else:
                            if str(sheet.cell(row=x+1, column=8).value) != "None":
                                description = description + "**Game " + str(x+1) + ":** | " + "Word: **" + str(sheet.cell(row=x+1, column=8).value) + "** | Guesses: **" + str(sheet.cell(row=x+1, column=7).value) + "** | " + winorloss + "\n"
                    embedVar = discord.Embed(title="Wordle Games : " + str(ctx.author), description=description, color=0x00ff00)
                    embedVar.set_footer(text= "Showing page 1 of " + str(int(pages)) + " (" + str(games) + " total games)")
                    await ctx.channel.send(embed=embedVar)
                else:
                    for x in range(10):
                        tempr = int(page) - 1
                        r = x + (int(tempr) * 10)
                        if sheet.cell(row=x+1, column=9).value == 0:
                            winorloss = "Loss"
                        elif sheet.cell(row=x+1, column=9).value == 1:
                            winorloss = "Win"
                        else:
                            winorloss = "Forfeit"
                        if x == 0:
                            if str(sheet.cell(row=r+1, column=8).value) != "None":
                                description = "**Game " + str(r+1) + ":** | " + "Word: **" + str(sheet.cell(row=r+1, column=8).value) + "** | Guesses: **" + str(sheet.cell(row=r+1, column=7).value) + "** | " + winorloss + "\n"
                            else:
                                await ctx.channel.send(str(ctx.author.mention) + " you havent played wordle before, do ``fstartwordle`` to start one")
                                return
                        else:
                            if str(sheet.cell(row=r+1, column=8).value) != "None":
                                description = description + "**Game " + str(r+1) + ":** | " + "Word: **" + str(sheet.cell(row=r+1, column=8).value) + "** | Guesses: **" + str(sheet.cell(row=r+1, column=7).value) + "** | " + winorloss + "\n"
                    embedVar = discord.Embed(title="Wordle Games : " + str(ctx.author), description=description, color=0x00ff00)
                    embedVar.set_footer(text= "Showing page " + str(int(page)) + " of " + str(int(pages)) + " (" + str(games) + " total games)")
                    await ctx.channel.send(embed=embedVar)

# ------------------------------------------------------- Show Game ------------------------------------------------------
    if contentWords[0].lower() == "fshowgame" or contentWords[0].lower() == "fsg":  
        game  = re.findall(r'[0-9]+', ctx.content)
        if len(ctx.content) < 1:
            await ctx.channel.send(str(ctx.author.mention) + " invalid command usages, do ``fshowgame`` ``(game number)``")
        else:
            game = game[0]
            game = int(game)
            workbook = load_workbook(filename="players/" + str(ctx.author.id) + "/wordles")
            sheet = workbook.active
            totalGames = len(sheet['A'])
            if totalGames < game:
                await ctx.channel.send(str(ctx.author.mention) + " invalid game number, do ``fgames`` ``*user*``, to see all games for the specified user")
            else:
                if game == 0:
                    await ctx.channel.send(str(ctx.author.mention) + " that game doesnt exist, do ``fshowgame`` ``(game number)``, to get games do ``fgames``")
                else:
                    for x in range(6):
                        if x == 0:
                            description = str(sheet.cell(row=game, column=x+1).value) + "\n"
                        else:
                            if str(sheet.cell(row=game, column=x+1).value) == "None":
                                description = description + "\n"
                            else:
                                description = description + str(sheet.cell(row=game, column=x+1).value) + "\n"
                    embedVar = discord.Embed(title="Wordle Game " +str(game), description=description, color=0x00ff00)
                    embedVar.set_footer(text=  "bold = correct letter, correct spot\nitalic = correct letter, wrong spot\nnothing = wrong letter, wrong spot")
                    await ctx.channel.send(embed=embedVar)
        
# ------------------------------------------------------- Displays the board ------------------------------------------------------
    if contentWords[0].lower() == "fboard" or contentWords[0].lower() == "fb":
        authorrow = idcheckf(ctx.author.id)
        workbook = load_workbook(filename="wordle/wordle")
        sheet = workbook.active
        alreadystarted = str(sheet.cell(row=authorrow, column=2).value)
        game = int(sheet.cell(row=authorrow, column=4).value) + int(sheet.cell(row=authorrow, column=5).value) + 1 
        workbook.close
        if alreadystarted == "0":
            await ctx.channel.send("You havent started wordle, type ``fstartwordle`` to begin.")
        else:
            workbook = load_workbook("players/" + str(ctx.author.id) + "/wordles")
            sheet = workbook.active
            for x in range(6):
                if x == 0:
                    description = str(sheet.cell(game, column=1).value) + "\n"
                else:
                    if str(sheet.cell(game, column=x+1).value) == "None":
                        description = description + "\n"
                    else:
                        description = description + str(sheet.cell(game, column=x+1).value) + "\n"
            embedVar = discord.Embed(title="Wordle : " + str(ctx.author), description=description, color=0x00ff00)
            embedVar.set_footer(text= "bold = correct letter, correct spot\nitalic = correct letter, wrong spot\nnothing = wrong letter, wrong spot")
            await ctx.channel.send(embed=embedVar)

# ------------------------------------------------------- User Info -------------------------------------------------------
    if contentWords[0].lower() == "fui" or contentWords[0].lower() == "fuserinfo":
        authorRow = idcheckf(ctx.author.id)
        workbook = load_workbook(filename="wordle/wordle.xlsx")
        sheet = workbook.active
        wordleswon = str(sheet.cell(row=authorRow, column=4).value)
        wordlesfailed = str(sheet.cell(row=authorRow, column=5).value)
        totalguesses = str(sheet.cell(row=authorRow, column=6).value)

        description = "Wordles Won: **" + wordleswon +"** \Wordles Failed: **" + wordlesfailed + "** \nTotal Guesses: **" + totalguesses
        embedVar = discord.Embed(title="View User Info", description="Showing statistics for " + str(ctx.author.mention) + "\n \n" + description, color=0x00ff00)
        await ctx.channel.send(embed=embedVar)

# ------------------------------------------------------- For testing purposes, resets wordle -------------------------------------------------------
    if contentWords[0].lower() == "fresetwordle" or contentWords[0].lower() == "frf" and ctx.author.id == adminid:
        authorrow = idcheckf(ctx.author.id)
        workbook = load_workbook(filename="wordle/wordle")
        sheet = workbook.active
        sheet["B" + str(authorrow)] = 0
        sheet["C" + str(authorrow)] = None
        sheet["D" + str(authorrow)] = 0
        sheet["E" + str(authorrow)] = 0
        sheet["F" + str(authorrow)] = 0
        workbook.save(filename="wordle/wordle")
        filename = "players/" + str(ctx.author.id) + "/wordles"
        os.remove(filename)
        wb = Workbook()
        wb.save("players/" + str(ctx.author.id) + "/wordles")
        await ctx.channel.send("your wordle data has been reset")

# ------------------------------------------------------- Resets Specific Person -------------------------------------------------------
    if contentWords[0].lower() == "fresetperson"  or contentWords[0].lower() == "frp" and ctx.author.id == adminid:
        userid  = re.findall(r'[0-9]+', ctx.content)
        if len(ctx.content) < 1:
            await ctx.channel.send(str(ctx.author.mention) + " invalid command usages, do ``fresetwordleperson`` ``(user id)``")
        else:
            workbook = load_workbook(filename="wordle/wordle")
            sheet = workbook.active
            max_row=sheet.max_row
            userid = userid[0]
            for i in range(1,max_row+1):
                cell_obj = sheet.cell(row=i, column=1)
                if str(cell_obj.value) == str(userid):
                    sheet["B" + str(userid)] = 0
                    sheet["C" + str(userid)] = None
                    sheet["D" + str(userid)] = 0
                    sheet["E" + str(userid)] = 0
                    sheet["F" + str(userid)] = 0
                    workbook.save(filename="wordle/wordle")
                    filename = "players/" + str(userid) + "/wordles"
                    os.remove(filename)
                    wb = Workbook()
                    wb.save(filename)
                    await ctx.channel.send(str(ctx.author.mention) + " " + str(userid) + "'s wordle has been reset")
                    break
                elif i == max_row and cell_obj.value != str(userid):
                    await ctx.channel.send(ctx.author.mention + " that user never played wordle before, or the ID is invalid")
                    break

# ------------------------------------------------------- Word Guess-------------------------------------------------------
    if contentWords[0].lower() == "fguess":
        authorrow = idcheckf(ctx.author.id)
        workbook = load_workbook(filename="wordle/wordle")
        sheet = workbook.active
        alreadystarted = str(sheet.cell(row=authorrow, column=2).value)
        word = str(sheet.cell(row=authorrow, column=3).value)
        game = int(sheet.cell(row=authorrow, column=4).value) + int(sheet.cell(row=authorrow, column=5).value)
        workbook.close
        wordLetters = list(word)
        fullWordGuess = ctx.content.lower().split()

# Checking for if the user has started a wordle
        if alreadystarted == "0":
            await ctx.channel.send("You havent started wordle, type ``fstartwordle`` to begin.")
            return

#Check for if there is a guess       
        if len(fullWordGuess) != 2:
            await ctx.channel.send("Invalid command usage, use ``fguess`` ``word``")
            return
        fullWordGuess.pop(0)
        wordGuess = list(fullWordGuess[0])
    
#check if guess is 5 letters long
        if len(wordGuess) != 5:
            await ctx.channel.send("Invalid word length, word must be 5 characters long")
            return

#Checking if word is a valid guess
        workbook = load_workbook(filename="wordle/words.xlsx")
        sheet = workbook.active
        max_row=sheet.max_row
        found = 0
        for x in range(max_row+1):
            if fullWordGuess[0] == sheet.cell(row=x+1, column=1).value:
                found = 1
                break
        if found != 1:
            await ctx.channel.send(fullWordGuess[0] + " is not a valid guess")
        else:
            workbook.close

# Check if the user has already started a game and load that game file  
            if alreadystarted == "1":
                game = game + 1
                workbook = load_workbook("players/" + str(ctx.author.id) + "/wordles")
                sheet = workbook.active
                if str(sheet.cell(row=game, column=7).value) == "None":
                    sheet["G" + str(game)] = 1
                    workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")
                    guessNumber =  1   
                else:
                    sheet["G" + str(game)]  = int(sheet.cell(row=game, column=7).value) + 1
                    workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")
                    guessNumber =  str(sheet.cell(row=game, column=7).value)

# Checking guess for correct letters, as well as storing guess in player's .xlsx file
                letterscorrect = 0
                wordGuessChecked = wordGuess
                for x in range(5):
                    if wordGuess[x] == wordLetters[x]:
                        wordGuessChecked[x] = "**" + str(wordGuess[x]) + "**"
                        letterscorrect = letterscorrect + 1  
                for x in range(5):
                    if wordGuess[x] in wordLetters:
                        lettercount = wordLetters.count(wordGuess[x])
                        correctLetterCount = wordGuessChecked.count("**" +str(wordGuess[x] + "**"))
                        if lettercount == correctLetterCount:
                            wordGuessChecked[x] = str(wordGuess[x])
                        else:
                            beans = wordGuessChecked.count("*" +str(wordGuess[x] + "*"))
                            if beans == 1:
                                wordGuessChecked[x] = str(wordGuess[x])
                            else:
                                wordGuessChecked[x] = "*" + str(wordGuess[x]) + "*"
                    else:
                        wordGuessChecked[x] = str(wordGuess[x])
                columnLetter = ["A", "B", "C", "D", "E", "F"]
                fullWordGuessChecked = " ".join(wordGuessChecked)
                workbook = load_workbook("players/" + str(ctx.author.id) + "/wordles")
                sheet = workbook.active     
                column = columnLetter[int(guessNumber) - 1]
                sheet[str(column) + str(game)] = str(fullWordGuessChecked)
                workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")

# if the user hasnt used all their guesses and the word isnt correct    
                if int(guessNumber) <= 6 and letterscorrect != 5:
                    for x in range(6):
                        if x == 0:
                            if guessNumber == 1:
                                description = str(sheet.cell(row=game, column=1).value) + "\n"
                            else:
                                description = str(sheet.cell(row=game, column=1).value) + "\n"
                        else:
                            if str(sheet.cell(row=game, column=x+1).value) == "None":
                                description = description + "\n"
                            else:
                                description = description + str(sheet.cell(row=game, column=x+1).value) + "\n"
                    embedVar = discord.Embed(title="Wordle : " + str(ctx.author), description=description, color=0x00ff00)
                    embedVar.set_footer(text=  "bold = correct letter, correct spot\nitalic = correct letter, wrong spot\nnothing = wrong letter, wrong spot")

# If the user has used all 6 guesses and the word has not guessed 
                    if int(guessNumber) == 6 and letterscorrect < 5:
                        await ctx.channel.send(str(ctx.author.mention) + " you failed the wordle. The word was: " + word, embed=embedVar)
                        sheet["H" + str(game)] = str(word)
                        sheet["G" + str(game)] = int(guessNumber)
                        sheet["I" + str(game)] = 0
                        workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")

                        workbook = load_workbook(filename="wordle/wordle")
                        sheet = workbook.active
                        sheet["B" + str(authorrow)] = 0
                        sheet["C" + str(authorrow)] = None  
                        sheet["E" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                        sheet["F" + str(authorrow)] = int(sheet.cell(row=authorrow, column=6).value) + int(guessNumber)
                        workbook.save(filename="wordle/wordle")

                    else:
                        await ctx.channel.send(embed=embedVar)

#If the user has guessed the word correctly in or under 6 guesses
                elif int(guessNumber) <= 6 and letterscorrect == 5:
                    for x in range(int(guessNumber)):
                        if x == 0:
                            description = str(sheet.cell(row=game, column=x+1).value) + "\n"
                        else: 
                            if guessNumber != 1:
                                if str(sheet.cell(row=game, column=x+1).value) == "None":
                                    description = description + "\n"
                                else:
                                    description = description + str(sheet.cell(row=game, column=x+1).value) + "\n"
                    for x in range(6 - int(guessNumber)):
                        description = description + "\n"
                    embedVar = discord.Embed(title="Wordle : " + str(ctx.author), description=description, color=0x00ff00)
                    embedVar.set_footer(text=  "bold = correct letter, correct spot\nitalic = correct letter, wrong spot\nnothing = wrong letter, wrong spot")
                    rarity = random.randrange(1, 1000)
                    if rarity <= 400: #40%
                        bug = "**common**"
                        column = 3
                    elif rarity >= 401 and rarity <= 600: #20%
                        bug = "**uncommon**"
                        column = 4
                    elif rarity >= 601 and rarity <= 750: #15%
                        bug = "**rare**" 
                        column = 5
                    elif rarity >= 751 and rarity <= 900: #15%
                        bug = "**epic**" 
                        column = 6
                    elif rarity >= 901 and rarity <= 1000: #10%
                        bug = "**legendary**"
                        column = 7

                    workbook = load_workbook(filename="tamagotchi/inventories.xlsx")
                    sheet = workbook.active
                    max_row=sheet.max_row
                    found = 0
                    for i in range(1,max_row+1):
                        cell_obj = sheet.cell(row=i, column=1)
                        if cell_obj.value == str(ctx.author.id):
                            row = i
                            found = 1
                            if sheet.cell(row=row, column=2).value != str(ctx.author.name):
                                sheet.cell(row=row, column=2).value = str(ctx.author.name)
                        elif i == max_row and cell_obj.value != str(ctx.author.id) and found != 1:
                            j = i + 1
                            row = j
                            sheet ["A" + str(j)] = str(ctx.author.id)
                            sheet ["B" + str(j)] = 0
                            sheet ["C" + str(j)] = 0
                            sheet ["D" + str(j)] = 0
                            sheet ["E" + str(j)] = 0
                            sheet ["F" + str(j)] = 0
                            if sheet.cell(row=row, column=2).value != str(ctx.author.name):
                                sheet.cell(row=row, column=2).value = str(ctx.author.name)
                            workbook.save(filename="tamagotchi/inventories.xlsx")
                    workbook.close
                    workbook = load_workbook(filename="tamagotchi/inventories.xlsx")
                    sheet = workbook.active
                    sheet.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value + 1
                    workbook.save(filename="tamagotchi/inventories.xlsx")
                    await ctx.channel.send("Congrats! " + str(ctx.author.mention) + " you guessed the wordle in " + str(guessNumber) + " guesses and won a " + str(bug) + " bug!", embed=embedVar)

#Save player file
                    workbook = load_workbook("players/" + str(ctx.author.id) + "/wordles")
                    sheet = workbook.active
                    sheet["H" + str(game)] = str(word)
                    sheet["G" + str(game)] = int(guessNumber)
                    sheet["I" + str(game)] = 1
                    workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")

#Save stats and reset for next game
                    workbook = load_workbook(filename="wordle/wordle")
                    sheet = workbook.active
                    sheet["B" + str(authorrow)] = 0
                    sheet["C" + str(authorrow)] = None
                    sheet["D" + str(authorrow)] = int(sheet.cell(row=authorrow, column=4).value) + 1
                    sheet["F" + str(authorrow)] = int(sheet.cell(row=authorrow, column=6).value) + int(guessNumber)
                    workbook.save(filename="wordle/wordle")

# if user hasnt started wordle
            else:
                await ctx.channel.send("You havent started wordle, type ``fstartwordle`` to begin.")

# ------------------------------------------------------- Starts Wordle -------------------------------------------------------
    if contentWords[0].lower() == "fstartwordle" or contentWords[0].lower() == "fsw":
        id = ctx.author.id
        authorrow = idcheckf(id)
        workbook = load_workbook(filename="wordle/wordle")
        sheet = workbook.active
        alreadystarted = str(sheet.cell(row=authorrow, column=2 ).value)
        workbook.close
        reply = 0
        if alreadystarted == "1":
            await ctx.channel.send("You already have a wordle started, to print the board type ``fboard``")
        else:
            message = await ctx.channel.send("Start Wordle?", components = [[Button(label="Yes", style="3", custom_id="button1" + str(ctx.author.id)), Button(label="No", style="4", custom_id="button2" + str(ctx.author.id))]])
            while True:
                try:
                    event = await bot.wait_for("button_click", timeout = 30.0)
                except asyncio.TimeoutError: 
                    if reply != 1:
                        await ctx.channel.send(f"{ctx.author.mention}, you didn't reply fast enough..") 
                        try:
                            await message.delete()
                        except discord.NotFound:
                            return
                        return
                else:
                    if event.user.id == ctx.author.id:
                        if event.component.id == "button1" + str(ctx.author.id):
                            reply = 1
                            try:
                                await message.delete()
                            except discord.NotFound:
                                return
                            workbook = load_workbook(filename="wordle/words.xlsx")
                            sheet = workbook.active   
                            tempvariable = False
                            x = 1
                            while tempvariable == False:
                                if sheet.cell(row=x, column=2).value == None:
                                    rowLength = x
                                    tempvariable = True
                                else:
                                    x = x + 1
                            ran = random.randrange(1, rowLength-1)
                            word = sheet.cell(row=int(ran), column=2).value
                            workbook.close
                            workbook = load_workbook(filename="wordle/wordle")
                            sheet = workbook.active
                            sheet["C" + str(authorrow)] = str(word)
                            sheet["B" + str(authorrow)] = 1
                            game = int(sheet.cell(row=authorrow, column=4).value) + int(sheet.cell(row=authorrow, column=5).value)
                            workbook.save(filename="wordle/wordle")
                            game = game + 1
                            if os.path.exists("players/" + str(ctx.author.id)) == False:
                                os.mkdir("players/" + str(ctx.author.id))
                                wb = Workbook()
                                wb.save("players/" + str(ctx.author.id) + "/wordles")
                            await ctx.channel.send(str(ctx.author.mention) + " wordle has been started, to guess a word do ``fguess`` ``word``, to see the board do fboard")
                        else:
                            reply = 1
                            await ctx.channel.send(content="Wordle was cancelled")
                            try:
                                await message.delete()
                            except discord.NotFound:
                                return
                            return
    
# ------------------------------------------------------- Ban Letter Or Word -------------------------------------------------------
    elif contentWords[0].lower() == "fbanword" or contentWords[0].lower() == "fbw" and ctx.author.id == adminid:
        workbook = load_workbook(filename="wordle/bannedlist.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        rowLength=len(sheet['A'])
        for x in range(rowLength+1):
            if word[1] == sheet.cell(row=x+1, column=1).value:
                await ctx.channel.send(str(word[1]) + " is already banned")
                break
            elif x == rowLength:
                sheet ["A" + str(rowLength+1)] = word[1]
                workbook.save(filename="wordle/bannedlist.xlsx")
                await ctx.channel.send(str(word[1]) + " added to the banned letter list")
        workbook.close

# ------------------------------------------------------- Unban Letter Or Word -------------------------------------------------------
    elif contentWords[0].lower() == "funban" or contentWords[0].lower() == "fub" and ctx.author.id == adminid:
        workbook = load_workbook(filename="wordle/bannedlist.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        rowLength=len(sheet['A'])
        for x in range(rowLength+1):
            if word[1] == sheet.cell(row=x+1, column=1).value:
                sheet.delete_rows(x+1)
                workbook.save(filename="wordle/bannedlist.xlsx")
                await ctx.channel.send(str(word[1]) + " has been unbanned")
                break
            elif x == rowLength:
                await ctx.channel.send(str(word[1]) + " isnt banned")
        workbook.close

# ------------------------------------------------------- Delete Word From Pool -------------------------------------------------------
    if contentWords[0].lower() == "fdeleteword" or contentWords[0].lower() == "fdw" and ctx.author.id == adminid:
        workbook = load_workbook(filename="wordle/words.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        tempvariable = False
        x = 1
        while tempvariable == False:
            if sheet.cell(row=x, column=2).value == None:
                rowLength = x
                tempvariable = True
            else:
                x = x + 1
        for x in range(rowLength+1):
            if word[1] == sheet.cell(row=x+1, column=2).value:
                sheet ["A" + str(x+1)] = None
                workbook.save(filename="words.xlsx")
                await ctx.channel.send(str(word[1]) + " has been removed from the word pool")
                break
            elif sheet.cell(row=x+1, column=2).value != None:
                if word[1] == sheet.cell(row=x+1, column=1).value:
                    sheet ["B" + str(x+1)] = None
                    workbook.save(filename="words.xlsx")
                    await ctx.channel.send(str(word[1]) + " has been removed from the word pool")
                    break
            elif x == rowLength:
                await ctx.channel.send(str(word[1]) + " isnt in the word list")
        workbook.close

# ------------------------------------------------------- Delete Guess From Pool -------------------------------------------------------
    if contentWords[0].lower() == "fdeleteguess" or contentWords[0].lower() == "fdg" and ctx.author.id == adminid:
        workbook = load_workbook(filename="wordle/words.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        rowLength=len(sheet['A'])
        for x in range(rowLength+1):
            if word[1] == sheet.cell(row=x+1, column=1).value:
                sheet ["A" + str(x+1)] = None
                workbook.save(filename="words.xlsx")
                await ctx.channel.send(str(word[1]) + " has been removed from the guess pool")
                break
            elif sheet.cell(row=x+1, column=1).value != None:
                if word[1] == sheet.cell(row=x+1, column=1).value:
                    sheet ["B" + str(x+1)] = None
                    workbook.save(filename="words.xlsx")
                    await ctx.channel.send(str(word[1]) + " has been removed from the guess pool")
                    break
            elif x == rowLength:
                await ctx.channel.send(str(word[1]) + " isnt in the guess list")
        workbook.close
    
# ------------------------------------------------------- Ban Letter or Word List -------------------------------------------------------
    if contentWords[0].lower() == "fnolist" or contentWords[0].lower() == "fnl":
        workbook = load_workbook(filename="wordle/bannedlist.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        rowLength=len(sheet['A'])
        x = 0
        ee = [] 
        for x in range(rowLength):
            ee.append(sheet.cell(row=x+1, column=1).value)
        ee.pop(0)
        temp = ', '.join(ee)
        await ctx.channel.send("The current banned letters and words are: " + temp)
        workbook.save(filename="wordle/bannedlist.xlsx")
        workbook.close

# ------------------------------------------------------- Forfeit -------------------------------------------------------
    if contentWords[0].lower() == "fforfeit" or contentWords[0].lower() == "ff":
        authorrow = idcheckf(ctx.author.id)
        workbook = load_workbook(filename="wordle/wordle")
        sheet = workbook.active
        alreadystarted = str(sheet.cell(row=authorrow, column=2).value)
        word = str(sheet.cell(row=authorrow, column=3).value)
        game = int(sheet.cell(row=authorrow, column=4).value) + int(sheet.cell(row=authorrow, column=5).value)
        if alreadystarted == "1":
            message = await ctx.channel.send("Forfeit Wordle?", components = [[Button(label="Yes", style="3", custom_id="button1" + str(ctx.author.id)), Button(label="No", style="4", custom_id="button2" + str(ctx.author.id))]])
            while True:
                try:
                    event = await bot.wait_for("button_click", timeout = 30.0)
                except asyncio.TimeoutError: 
                    if reply != 1:
                        await ctx.channel.send(f"{ctx.author.mention}, you didn't reply fast enough..") 
                        try:
                            await message.delete()
                        except discord.NotFound:
                            return
                        return
                else:
                    if event.user.id == ctx.author.id:
                        if event.component.id == "button1" + str(ctx.author.id):
                            reply = 1
                            try:
                                await message.delete()  
                            except discord.NotFound:
                                return
                            workbook.close
                            workbook = load_workbook("players/" + str(ctx.author.id) + "/wordles")
                            sheet = workbook.active
                            game = game + 1
                            found = 0
                            for x in range(6):
                                if sheet.cell(row=game, column=x+1).value == None and found == 0:
                                    guessNumber  = x
                                    found = 1
                            
                            if guessNumber == 0:
                                workbook.close
                                workbook = load_workbook(filename="wordle/wordle")
                                sheet = workbook.active
                                sheet["B" + str(authorrow)] = 0
                                sheet["C" + str(authorrow)] = None
                                workbook.save(filename="wordle/wordle")
                            else:
                                sheet["H" + str(game)] = str(word)
                                sheet["G" + str(game)] = int(guessNumber)
                                sheet["I" + str(game)] = 3
                                workbook.save(filename="players/" + str(ctx.author.id) + "/wordles")

                                workbook = load_workbook(filename="wordle/wordle")
                                sheet = workbook.active
                                sheet["B" + str(authorrow)] = 0
                                sheet["C" + str(authorrow)] = None
                                sheet["E" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                sheet["F" + str(authorrow)] = int(sheet.cell(row=authorrow, column=6).value) + 6
                                workbook.save(filename="wordle/wordle")
                            await ctx.channel.send(str(ctx.author.mention) + " your wordle has been forfeited, the word was: **" + str(word) + "**")                                   
        else:
            await ctx.channel.send("You havent started a wordle, do ``fstartwordle`` to start one")

# ------------------------------------------------------- Shows Person's Word -------------------------------------------------------
    if contentWords[0].lower() == "fword" and ctx.author.id == adminid:
        userid  = re.findall(r'[0-9]+', ctx.content)
        if len(ctx.content) < 1:
            await ctx.channel.send(str(ctx.author.mention) + " invalid command usages, do ``fword`` ``(user id)``")
        else:
            workbook = load_workbook(filename="wordle/wordle")
            sheet = workbook.active
            max_row=sheet.max_row
            userid = userid[0]
            for i in range(1,max_row+1):
                cell_obj = sheet.cell(row=i, column=1)
                if str(cell_obj.value) == str(userid):
                    if sheet.cell(row=i, column=2).value == 1:
                        await ctx.channel.send(sheet.cell(row=i, column=3).value)
                        break
                    else:
                        await ctx.channel.send(ctx.author.mention + " that user hasnt started a wordle")
                        break
                elif i == max_row and cell_obj.value != str(userid):
                    await ctx.channel.send(ctx.author.mention + " that user never played wordle before, or the ID is invalid")
                    break

# ------------------------------------------------------- Add Word -------------------------------------------------------
    if contentWords[0].lower() == "faw" or contentWords[0].lower() == "frogeaddword":
        workbook = load_workbook(filename="wordle/bannedlist.xlsx")
        sheet = workbook.active
        word = ctx.content.split()
        theword = word[1].lower()
        rowLength=len(sheet['B'])
        bannedword = 0
        for x in range(rowLength+1):
            if str(sheet.cell(row=x+1, column=1).value) in str(ctx.content):
                bannedword = 1
                break
        workbook.close
        workbook = load_workbook(filename="wordle/words.xlsx")
        sheet = workbook.active
        if len(word) == 1:
            await ctx.channel.send("Please specify a word")
        elif len(theword) != 5:
            await ctx.channel.send("Invalid word length")
        elif theword.isalpha() == False:
            await ctx.channel.send("Invalid word")
        elif bannedword ==1:
            await ctx.channel.send("You have used a banned character or " + str(theword) + " is a banned word")
        else:
            tempvariable = False
            x = 1
            while tempvariable == False:
                if sheet.cell(row=x, column=2).value == None:
                    rowLength = x
                    tempvariable = True
                else:
                    x = x + 1
            max_row = len(sheet['A'])
            found = 0 
            for x in range(max_row+1):
                if theword == str(sheet.cell(row=x+1, column=1).value):
                    found =1
            for j in range(rowLength+1):
                if theword == str(sheet.cell(row=j+1, column=2).value):
                    if found == 1:
                        await ctx.channel.send(theword + " is already in the word pool")
                        return
                    elif found == 0:
                        sheet ["A" + str(max_row+1)] = theword
                        workbook.save(filename="wordle/words.xlsx")
                        await ctx.channel.send(theword + " has been added to the guess pool")
                        return
                elif j == rowLength:
                    if found == 1:
                        sheet ["B" + str(rowLength+1)] = theword
                        workbook.save(filename="wordle/words.xlsx")
                        await ctx.channel.send(theword + " has been added to the word pool")
                        return
                    elif found == 0:
                        sheet ["A" + str(max_row+1)] = theword
                        sheet ["B" + str(rowLength+1)] = theword
                        workbook.save(filename="wordle/words.xlsx")
                        await ctx.channel.send(theword + " has been added to the word and guess pool")
                        return


# -------------------------------------------------------------------------------------------------------------- Functions --------------------------------------------------------------------------------------------------------------

# ------------------------------------------------------- ID check, input user id, returns row of author in wordle-------------------------------------------------------
def idcheckf(id):
    workbook = load_workbook(filename="wordle/wordle")
    sheet = workbook.active
    max_row=sheet.max_row
    found = 0
    for i in range(1,max_row+1):
        cell_obj = sheet.cell(row=i, column=1)
        if cell_obj.value == str(id):
            row = i
            found = 1
        elif i == max_row and cell_obj.value != str(id) and found != 1:
            j = i + 1
            sheet ["A" + str(j)] = str(id)
            sheet ["B" + str(j)] = 0
            sheet ["D" + str(j)] = 0
            sheet ["E" + str(j)] = 0
            sheet ["F" + str(j)] = 0
            workbook.save(filename="wordle/wordle")
            row = j
    workbook.close
    return row


# ------------------------------------------------------- Error checking -------------------------------------------------------
@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, CommandNotFound):
        return

bot.run(token)